import openpyxl as xl
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from docxtpl import DocxTemplate 
import time
import sys
import os

base_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()
#base_dir = getattr(sys, '_MEIPASS', Path(__file__).resolve().parent)

word_template_path = base_dir / "VALIANT TMS NCR TAG.docx"
word_template_path1 = base_dir / "Email template.docx"
excel_template_path = base_dir / "ShipRequest.xlsx"
excel_path = base_dir / "JeremyNCR.xlsx"
output_dir = base_dir / "OUTPUT"

# # Specify the relative paths to the template files
# word_template_path = os.path.join(base_dir, "VALIANT TMS NCR TAG.docx")
# word_template_path1 = os.path.join(base_dir, "Email template.docx")
# excel_template_path = os.path.join(base_dir, "ShipRequest.xlsx")

# # Specify the absolute paths to the input and output files
# excel_path = os.path.join(base_dir, "JeremyNCR.xlsx")
# output_dir = os.path.join(base_dir, "OUTPUT")


# Create output folder for the word documents
output_dir.mkdir(exist_ok=True)
#Path(output_dir).mkdir(exist_ok=True)

# Convert Excel sheet to pandas dataframe
df = pd.read_excel(excel_path, sheet_name="SAPUI5 Export")
dc = pd.read_excel(excel_path, sheet_name="SHEET1")

# Prompt user to enter vendor name (without requiring "NCR-")
vendor_input = input("Enter NCR number: NCR- ")

# Add "NCR-" prefix to the user input
NCR = "NCR-" + vendor_input

# Prompt user for the desired action
print("Select an option:")
print("1. Generate Email")
print("2. Generate NCR tag")
print("3. Generate Shipping Request")

# Get user's choice
option = input("Enter option number: ")

# Start the timer
start_time = time.time()

# Create a dictionary to store the vendor information from the second sheet
vendor_info = {}
email_dict = {}
for record in dc.to_dict(orient="records"):
    vendor_info[record['Vendor']] = record
    email_dict[record['Vendor']] = record['Email'] 

# Load the workbook
wb = load_workbook(excel_template_path)
ws = wb.active

# Initialize the output file name
output_file_name = None

if option == "1":
    # Generate Email
    for record in df.to_dict(orient="records"):
        if record["Control_Num"].startswith(NCR):
            doc = DocxTemplate(word_template_path1)
            doc.render(record)
            output_path = output_dir / f"{record['External_provider']}-Email Temp.docx"
            doc.save(output_path)                    
    print("Email generated successfully.")

elif option == "2":
    # Generate NCR tag
    for record in df.to_dict(orient="records"):
        if record["Control_Num"].startswith(NCR):
            # ===========================Generate NCR TAG===========================
            doc = DocxTemplate(word_template_path)
            doc.render(record)
            output_path = output_dir / f"{record['External_provider']}-NCR Tag.docx"
            doc.save(output_path)
            pass
    print("NCR tag generated successfully.")
# Validate and perform the selected action

elif option == "3":
    # Generate Shipping Request
    for record in df.to_dict(orient="records"):
        if record["Control_Num"].startswith(NCR):
            # Job number
            ws['C3'] = record['Job_No']
            # Company Name
            ws['B7'] = record['External_provider']
            ws['K16'] = record['External_provider']
            # PO
            ws['A16'] = record['PO']
            # Qty
            ws['B16'] = record['Quantity']
            # Part number
            ws['C16'] = record['Part_Number']
            # Project Manager
            ws['C4'] = record['PM']
            
            # Retrieve the matching vendor information from the dictionary
            vendor = record['External_provider']
            if vendor in vendor_info:
                dc_record = vendor_info[vendor]
                # Adress
                # Street
                ws['B8'] = dc_record['Street']
                # City
                ws['B9'] = dc_record['City']
                # State
                ws['E9'] = dc_record['State']
                # Country
                ws['B10'] = dc_record['Country']
                ws['I16'] = dc_record['Country']
                # Postal Code
                ws['E10'] = dc_record['Postalcode']
                # Phone
                ws['E11'] = dc_record['Phone']
                
            # Set the output file name if it hasn't been set already
            if output_file_name is None:
                output_file_name = f"{record['External_provider']}-Shipping Request.xlsx"
            pass
    print("Shipping Request generated successfully.")
else:
    print("Invalid option selected.")


# Check if an output file name was set
if output_file_name is not None:
    # Define the output file path with the respective Vendor/External Provider name
    output_path = output_dir / output_file_name

    # Save the workbook
    wb.save(output_path)

    # Calculate and print the execution time
    execution_time = time.time() - start_time
    print(f"Execution time: {execution_time} seconds")
else:
    print("No matching records found.")